import os
from fillpdf import fillpdfs
from openpyxl import load_workbook


def check_input():
    for in_file in os.listdir(os.getcwd()):
        if in_file.endswith(".xlsx"):
            return 1
    return 0


def check_output():
    for out_file in os.listdir(os.getcwd()):
        if out_file.endswith(".pdf"):
            return 1
    return 0


def error_message(flag):
    if flag == 1:
        error_file = open("НЕТ ЭКСЕЛЬ ФАЙЛА.txt", "w+")
        error_file.write("Отсутствует эксель файл для расчета данных и заполнения в пдф!\n ")
        error_file.write("Поместите недостающий файл в папку к исполняемому файлу(count.exe) и "
                         "запустите программу заново \n")
        error_file.write("Спасибо ;)")
        error_file.close()
    if flag == 2:
        error_file = open("НЕТ ПДФ ФАЙЛА.txt", "w+")
        error_file.write("Отсутствует шаблонный пдф файл для вывода корректного результата!\n ")
        error_file.write("Поместите шаблон пдф файла в папку к исполняемому файлу(count.exe) и "
                         "запустите программу заново \n")
        error_file.write("Спасибо ;)")
        error_file.close()
    if flag == 3:
        error_file = open("НЕТ ЭКСЕЛЬ И ПДФ ФАЙЛОВ.txt", "w+")
        error_file.write("Отсутствует эксель файл и шаблон пдф\n ")
        error_file.write("Поместите недостающие файлы в папку к исполняемому файлу(count.exe) и "
                         "запустите программу заново \n")
        error_file.write("Спасибо ;)")
        error_file.close()
    if flag == 4:
        return 1


def errors():
    if check_output() == 0 and check_input() == 0:
        error_message(3)
        return 1
    if check_input() == 0:
        error_message(1)
        return 1
    if check_output() == 0:
        error_message(2)
        return 1
    return 0


def read_xlsx_file():
    wb = load_workbook('file.xlsx', data_only=True)
    ws = wb.active
    data_1 = ws['C26'].value
    data_2 = ws['D26'].value
    data = [data_1, data_2]
    return data


def write_in_pdf_file(exel_data):
    dict_data = dict(zip(fillpdfs.get_form_fields("blank.pdf")))
    fillpdfs.write_fillable_pdf('blank.pdf', 'new_blank.pdf', dict_data)
    return 1


def program():
    if errors():
        return 0
    exel_data = read_xlsx_file()
    print(exel_data)
    write_in_pdf_file(exel_data)


program()
